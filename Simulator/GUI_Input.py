import sys, time
from PyQt4 import QtGui
from PyQt4.QtCore import Qt
import Simulate as sim
#from GUI_Output import Output
from GUI_Output2 import Output
import ast
from random import randint

from openpyxl import load_workbook

# N = 3
# # Total goods
# M = 1
# # Number of goods that are perishable
# M_perishable = 0
# perish_factor = 0
# # stable at prodcution_time = M * perish_factor
# production_time = 0
# value = 1

labels = ['Perish Period', 'Production Delay', 'Nominal Value']

class Input(QtGui.QMainWindow):
    
    def __init__(self):
        super(Input, self).__init__()

        self.initUI()

    def initUI(self):
        # Menu bar
        self.menubar = QtGui.QMenuBar()
        
        fileMenu = self.menubar.addMenu('&File')
        fileMenu.addAction('Load', lambda: self.load_data())

        self.setMenuBar(self.menubar)

        # Frames
        self.gui = GUI()
        self.setCentralWidget(self.gui)
        
        # Window size

        self.setGeometry(0, 50, 300, 475)
        self.setWindowTitle('The Giving Game - Input')     
        self.show()

    def load_data(self):
        fname = QtGui.QFileDialog.getOpenFileName(self, 'Open file', 
        '/Users/julianruger/Informatica/Afstudeerproject/Afstudeerproject/Simulator/results')
        
        # if fname == 'input.csv':
        #     #self.saveData(fname)
        #     print(fname)
        with open(fname) as f:
            content = f.readlines()
            N = int(content[0])
            M = int(content[1])
            if content[2].strip() == 'False':
                parallel = False
            else:
                parallel = True
            selectionrule = int(content[3])
            print(content[4])
            s = str(content[4].strip())
            print(s[1:-1])
            goods_list = ast.literal_eval(s[1:-1])
            #goods_list = [n.strip() for n in goods_list]
            print(goods_list[1][1])
            #print(goods_list[1][1])
            for x in goods_list:
                for y in x:
                    y = int(y)
            self.gui.load_simulation(N, M, goods_list, 0, 0, 0, 1, parallel, selectionrule)

    def closeEvent(self,event):
        result = QtGui.QMessageBox.question(self,
                      "Confirm Exit...",
                      "Are you sure you want to exit ?",
                      QtGui.QMessageBox.Yes| QtGui.QMessageBox.No)
        event.ignore()

        if result == QtGui.QMessageBox.Yes:
            event.accept()


class GUI(QtGui.QWidget):
    
    def __init__(self):
        super(GUI, self).__init__()
        
        # Total agents
        self.N = 10
        # Total goods
        self.M = 3
        # Number of goods that are perishable
        self.M_perishable = 0
        self.perish_period = 0
        # stable at prodcution_time = M * perish_period
        self.production_delay = 0
        self.value = 1

        # 0 = random rule
        # 1 = balance rule
        # 2 = goodwill rule
        self.selection_rule = 0

        self.parallel = False

        # [[perish_period, production_delay, nominal_value]]
        self.goods_list = []
        self.start_agents = []

        # xlsx variables
        self.like_factors = []
        self.balance = []
        self.nominal_values = []

        self.initUI()
        

    def initUI(self):

        hbox = QtGui.QHBoxLayout()

        screen = QtGui.QDesktopWidget().availableGeometry()
        #print('window size:', screen)
        # Agents
        self.lbl_nr_agents = QtGui.QLabel('Number of Agents')
        self.nr_agents = QtGui.QSpinBox()
        self.nr_agents.setMaximum(10000)

        self.nr_agents.valueChanged.connect(self.setAgents)

        # Goods
        self.lbl_nr_goods = QtGui.QLabel('Number of Goods')
        self.nr_goods = QtGui.QSpinBox()
        self.nr_goods.setMaximum(10000)

        self.nr_goods.valueChanged.connect(self.setGoods)

        # Table widget for the input of goods variables
        self.layout = QtGui.QGridLayout()
        self.lbl_input_table = QtGui.QLabel('Create goods')
        self.input_table = QtGui.QTableWidget()
        #self.input_table.cellChanged.connect(self.cellItemChanged)

        # Like factor
        self.load_likefactors = QtGui.QPushButton('Add like factors')
        self.load_likefactors.clicked.connect(self.setLikeFactors)
        self.lbl_likefactors = QtGui.QLabel('')

        self.remove_likefactors = QtGui.QPushButton('X')
        self.remove_likefactors.clicked.connect(self.removeLikeFactors)
        self.remove_likefactors.setFixedSize(25,25)

        hbox_likefactors = QtGui.QHBoxLayout()
        hbox_likefactors.addWidget(self.lbl_likefactors)
        hbox_likefactors.addWidget(self.remove_likefactors)
        #hbox_likefactors.addStretch(10)

        # Balance
        self.load_balance = QtGui.QPushButton('Add balance')
        self.load_balance.clicked.connect(self.setBalance)
        self.lbl_balance = QtGui.QLabel('')

        self.remove_balance = QtGui.QPushButton('X')
        self.remove_balance.clicked.connect(self.removeBalance)
        self.remove_balance.setFixedSize(25,25)

        hbox_balance = QtGui.QHBoxLayout()
        hbox_balance.addWidget(self.lbl_balance)
        hbox_balance.addWidget(self.remove_balance)

        # Nominal values
        self.load_nominalvalues = QtGui.QPushButton('Load nominal values')
        self.load_nominalvalues.clicked.connect(self.setNominalValues)

        self.lbl_nominalvalues = QtGui.QLabel('')

        self.remove_nominalvalues = QtGui.QPushButton('X')
        self.remove_nominalvalues.clicked.connect(self.removeNominalValues)
        self.remove_nominalvalues.setFixedSize(25,25)

        hbox_nominalvalues = QtGui.QHBoxLayout()
        hbox_nominalvalues.addWidget(self.lbl_nominalvalues)
        hbox_nominalvalues.addWidget(self.remove_nominalvalues)
        # Selection rules
        self.lbl_selection_rule = QtGui.QLabel('Choose selection rule')
        self.selection_rules = QtGui.QComboBox() 
        self.selection_rules.addItem('Random rule')
        self.selection_rules.addItem('Balance rule')
        self.selection_rules.addItem('Goodwill rule')


        self.selection_rules.activated[str].connect(self.setSelectionrule) 

        # Simulation type
        self.lbl_simulation_type = QtGui.QLabel('Simulation type')

        self.parallel = QtGui.QRadioButton('Parallel')
        self.onebyone = QtGui.QRadioButton('One by one')

        self.parallel.toggled.connect(self.setParallel)
        self.onebyone.toggled.connect(self.setOnebyOne)


        hbox.addWidget(self.parallel)
        hbox.addWidget(self.onebyone)

        # Start button
        self.start = QtGui.QPushButton('Simulate')
        self.start.setGeometry(100, 320, 100, 50)
        self.start.clicked.connect(self.startSimulation)

        self.layout.addWidget(self.lbl_nr_agents, 0, 0)
        self.layout.addWidget(self.nr_agents, 1, 0)
        self.layout.addWidget(self.lbl_nr_goods, 2, 0)
        self.layout.addWidget(self.nr_goods, 3, 0)
        self.layout.addWidget(self.lbl_input_table, 4, 0) 
        self.layout.addWidget(self.input_table, 5, 0)
        self.layout.addWidget(self.load_likefactors, 6, 0) 
        self.layout.addLayout(hbox_likefactors, 7, 0) 
        self.layout.addWidget(self.load_balance, 8, 0) 
        self.layout.addLayout(hbox_balance, 9, 0) 
        self.layout.addWidget(self.load_nominalvalues, 10, 0)
        self.layout.addLayout(hbox_nominalvalues, 11, 0)   
        self.layout.addWidget(self.lbl_selection_rule, 12, 0)
        self.layout.addWidget(self.selection_rules, 13, 0)
        self.layout.addWidget(self.lbl_simulation_type, 14, 0)
        self.layout.addLayout(hbox, 15, 0)
        self.layout.addWidget(self.start, 16, 0)

        self.setLayout(self.layout)

        # self.setGeometry(0, 50, 300, 475)
        # self.setWindowTitle('The Giving Game - Input') 
        # self.show()

    def setAgents(self, value): 
        self.N = self.nr_agents.value()
        print('Number of agents: ', self.N)

    def setGoods(self, value): 
        self.M = self.nr_goods.value()

        self.setInputTable(self.M)

        print('Total goods: ', self.M)

    def setPerishable(self, value):
        self.M_perishable = self.nr_perishable.value()
        if self.M_perishable > self.M:
            print("Too many perishable goods, number of perishable goods set to 0")
            self.M_perishable = 0
        else:
            print('Number of perishable goods:', self.M_perishable)

    def setPerishPeriod(self, value): 
        if self.M_perishable == 0:
            print('No goods are perishable, perish factor set to 0')
            self.perish_period = 0
        else:
            self.perish_period = self.perish.value()
        print('Perisch factor: ', self.perish_period)

    def setProductionTime(self, value): 
        self.production_delay = self.p_time.value()
        print('Production time: ', self.production_delay)

    def setValue(self, value): 
        self.value = self.good_value.value()
        print('Value of the goods: ', self.value)

    def setInputTable(self, M):
        goods_labels = []
        self.goods_list = []
        # Set number of rows
        self.input_table.setRowCount(M)
        # Set number of columns
        self.input_table.setColumnCount(4)
        # Set column headers
        self.input_table.setHorizontalHeaderLabels(labels)

        # Each row is a good, create the row headers and set the default values
        for i in range(M):
            lbl = 'Good_' + str(i)
            goods_labels.append(lbl)

            
            
            # Set default perish period
            item = QtGui.QTableWidgetItem()
            item.setData(Qt.EditRole, 0)
            self.input_table.setItem(i, 0, item)
            # Set default production delay
            # self.input_table.setItem(i, 1, QtGui.QTableWidgetItem('0'))
            item = QtGui.QTableWidgetItem()
            item.setData(Qt.EditRole, 0)
            self.input_table.setItem(i, 1, item)
            # Set default nominal value
            item = QtGui.QTableWidgetItem()
            item.setData(Qt.EditRole, 1)
            self.input_table.setItem(i, 2, item)
            # Set start agent
            item = QtGui.QTableWidgetItem()
            item.setData(Qt.EditRole, -1)
            self.input_table.setItem(i, 3, item)

            self.goods_list.append([0, 0, 1])

        # Set row headers
        self.input_table.setVerticalHeaderLabels(goods_labels)

    def setSelectionrule(self, text):
        if text == 'Random rule':
            self.selection_rule = 0
        elif text == 'Balance rule':
            self.selection_rule = 1
        elif text == 'Goodwill rule':
            self.selection_rule = 2

    def setParallel(self):
        self.parallel = True

    def setOnebyOne(self):
        self.parallel = False
    # @cellItemChanged(int, int)
    # def cellItemChanged(self, row, column):
    #     print(self.input_table.item(row, column))
    #     #self.goods_list[row][column] = self.input_table.item(row, column)
    def setGoodsList(self):
        rows = self.input_table.rowCount()
        columns = self.input_table.columnCount()
        self.start_agents = []
        # random.seed()
        for i in range(rows):
            for j in range(columns):
                if j == columns-1:
                    if self.input_table.item(i, j).data(Qt.DisplayRole) < 0 or self.input_table.item(i, j).data(Qt.DisplayRole) > self.N-1:
                        self.start_agents.append(randint(0, self.N-1))
                    else:
                        self.start_agents.append(self.input_table.item(i, j).data(Qt.DisplayRole))
                    print(self.input_table.item(i, j).data(Qt.DisplayRole))
                else:
                    self.goods_list[i][j] = self.input_table.item(i, j).data(Qt.DisplayRole)

    def setLikeFactors(self):
        fname = QtGui.QFileDialog.getOpenFileName(self, 'Open file', 
        '/Users/julianruger/Informatica/Afstudeerproject/Afstudeerproject/Simulator')

        self.like_factors = []
        if fname != '':
            name = fname.split('/')
            self.lbl_likefactors.setText(name[len(name) - 1])

            wb2 = load_workbook(fname)
            first_sheet = wb2.get_sheet_names()[0]
            worksheet = wb2.get_sheet_by_name(first_sheet)
            for row in worksheet.iter_rows():
                print(row)
                cells = []
                for cell in row:
                    print(cell.value)
                    if cell.value:
                        cells.append(float(cell.value))
                    else:
                        cells.append(-1.0)
                self.like_factors.append(cells)
            print(self.like_factors)

    def removeLikeFactors(self):
        self.lbl_likefactors.setText('')
        self.like_factors = []

    def setBalance(self):
        fname = QtGui.QFileDialog.getOpenFileName(self, 'Open file', 
        '/Users/julianruger/Informatica/Afstudeerproject/Afstudeerproject/Simulator')

        self.balance = []
        if fname != '':
            name = fname.split('/')
            self.lbl_balance.setText(name[len(name) - 1])

            wb2 = load_workbook(fname)
            first_sheet = wb2.get_sheet_names()[0]
            worksheet = wb2.get_sheet_by_name(first_sheet)
            for row in worksheet.iter_rows():
                print(row)
                cells = []
                for cell in row:
                    print(cell.value)
                    if cell.value:
                        cells.append(float(cell.value))
                    elif cell.value == 0:
                        cells.append(0.0)
                    else:
                        cells.append(None)
                self.balance.append(cells)
            print(self.balance)

            for i in range(len(self.balance)):
                for j in range(len(self.balance[1])):
                    if i != j and self.balance[i][j] != None:
                        if self.balance[i][j] != 0.0:
                            self.balance[j][i] = -self.balance[i][j]
                        elif self.balance[i][j] == 0.0:
                            self.balance[j][i] = self.balance[i][j]
                    if i == j:
                        self.balance[i][j] = 0.0
            print(self.balance)

    def removeBalance(self):
        self.lbl_balance.setText('')
        self.balance = []


    def setNominalValues(self):
        fname = QtGui.QFileDialog.getOpenFileName(self, 'Open file', 
        '/Users/julianruger/Informatica/Afstudeerproject/Afstudeerproject/Simulator')

        self.nominal_values = []
        if fname != '':
            name = fname.split('/')
            self.lbl_nominalvalues.setText(name[len(name) - 1])

            wb2 = load_workbook(fname)
            first_sheet = wb2.get_sheet_names()[0]
            worksheet = wb2.get_sheet_by_name(first_sheet)
            for row in worksheet.iter_rows():
                print(row)
                cells = []
                for cell in row:
                    print(cell.value)
                    if cell.value:
                        cells.append(float(cell.value))
                    else:
                        cells.append(randint(1, 5))
                self.nominal_values.append(cells)
            print(self.nominal_values)


    def removeNominalValues(self):
        self.lbl_nominalvalues.setText('')
        self.nominal_values = []

    def startSimulation(self):
        if self.check_parameters():
            self.setGoodsList()
            env = sim.create_enviroment(self.N, self.M, self.goods_list, self.M_perishable, self.perish_period, self.production_delay, self.value, self.parallel, self.selection_rule, self.like_factors, self.balance, self.nominal_values, self.start_agents)
            self.output = Output(env)
            env.output = self.output
            sim.start_simulation(self.N, self.M, self.goods_list, self.M_perishable, self.perish_period, self.production_delay, self.value, self.output, env, self.selection_rule)

    def load_simulation(self, N, M, goods_list, M_perishable, perish_period, production_delay, value, parallel, selection_rule):
        env = sim.create_enviroment(N, M, goods_list, M_perishable, perish_period, production_delay, value, parallel, selection_rule)
        output = Output(env)
        env.output = output
        sim.start_simulation(N, M, goods_list, M_perishable, perish_period, production_delay, value, output, env, selection_rule)

    def check_parameters(self):
        correct = True

        if self.N <= 0:
            correct = False
        if self.M <= 0:
            correct = False
        if self.like_factors and (len(self.like_factors) != self.N or len(self.like_factors[0]) != self.N):
            correct = False
        if self.balance and (len(self.balance) != self.N or len(self.balance[0]) != self.N):
            correct = False
        if self.nominal_values and (len(self.nominal_values) != self.N or len(self.nominal_values[0]) != self.M):
            correct = False

        if not correct:
            result = QtGui.QMessageBox.warning(self,
                      "Wrong parameter",
                      "Wrong parameters!",
                      QtGui.QMessageBox.Ok)

        return correct

    def closeEvent(self,event):
        result = QtGui.QMessageBox.question(self,
                      "Confirm Exit...",
                      "Are you sure you want to exit ?",
                      QtGui.QMessageBox.Yes| QtGui.QMessageBox.No)
        event.ignore()

        if result == QtGui.QMessageBox.Yes:
            event.accept()
     
        
def main():
    
    app = QtGui.QApplication(sys.argv)
    ex = Input()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main() 